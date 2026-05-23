"""Génère deux fichiers Excel d'import pour le premier semestre :
   - import_TIC_S1.xlsx        (département TIC)
   - import_CHIMIE_S1.xlsx     (département Chimie Appliquée)

Format attendu par ChronoFS (matieres/views.py:111) :
   filiere_code | niveau | code_ue | intitule | enseignant | type_seance | volume_horaire
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COULEUR_EN_TETE = "0D1520"
COULEUR_TEXTE = "C9A450"
COULEUR_LIGNE_PAIRE = "F7F4ED"

ENTETES = [
    "filiere_code",
    "niveau (L1/L2/L3/M1/M2)",
    "code_ue",
    "intitule",
    "enseignant",
    "type_seance (CM/TD/TP)",
    "volume_horaire (h)",
]
LARGEURS = [14, 22, 14, 48, 32, 22, 18]

# ── Premier semestre TIC ─────────────────────────────────────────────
MATIERES_TIC = [
    # TIC L1 — Semestre 1
    ("TIC", "L1", "TIC111", "Algorithmique & structures de données",       "Dr. Nyabeye Pierre",     "CM", 2.5),
    ("TIC", "L1", "TIC112", "Architecture des ordinateurs",                "Dr. Eloundou Marc",      "CM", 2.5),
    ("TIC", "L1", "TIC113", "Mathématiques pour l'informatique",           "Pr. Mbarga Joseph",      "CM", 2.5),
    ("TIC", "L1", "TIC114", "Introduction aux systèmes d'information",     "Dr. Owona Sylvie",       "CM", 2.5),
    ("TIC", "L1", "TIC115", "Programmation procédurale (C)",               "Dr. Nyabeye Pierre",     "TP", 2.5),
    ("TIC", "L1", "TIC116", "Anglais technique I",                         "Mme. Edzimbi Claire",    "TD", 2.5),

    # TIC L2 — Semestre 3
    ("TIC", "L2", "TIC211", "Programmation orientée objet (Java)",         "Dr. Ngono Albert",       "CM", 2.5),
    ("TIC", "L2", "TIC212", "Bases de données relationnelles",             "Pr. Mboumboumbo Jean",   "CM", 2.5),
    ("TIC", "L2", "TIC213", "Systèmes d'exploitation",                     "Dr. Eloundou Marc",      "CM", 2.5),
    ("TIC", "L2", "TIC214", "Réseaux informatiques I",                     "Dr. Atangana Paul",      "CM", 2.5),
    ("TIC", "L2", "TIC215", "Méthodes numériques",                         "Pr. Mbarga Joseph",      "CM", 2.5),
    ("TIC", "L2", "TIC216", "TP Bases de données",                         "Pr. Mboumboumbo Jean",   "TP", 2.5),

    # TIC L3 — Semestre 5
    ("TIC", "L3", "TIC311", "Analyse & Conception de SI (UML)",            "Pr. Mboumboumbo Jean",   "CM", 2.5),
    ("TIC", "L3", "TIC312", "Développement web full-stack",                "Dr. Ngono Albert",       "CM", 2.5),
    ("TIC", "L3", "TIC313", "Sécurité informatique",                       "Dr. Atangana Paul",      "CM", 2.5),
    ("TIC", "L3", "TIC314", "Intelligence artificielle",                   "Pr. Mbarga Joseph",      "CM", 2.5),
    ("TIC", "L3", "TIC315", "Génie logiciel & gestion de projet",          "Dr. Owona Sylvie",       "TD", 2.5),
    ("TIC", "L3", "TIC316", "TP Développement web",                        "Dr. Ngono Albert",       "TP", 2.5),
]

# ── Premier semestre Chimie Appliquée ────────────────────────────────
MATIERES_CHIM = [
    # CHIM L1 — Semestre 1
    ("CHIM", "L1", "CHM111", "Chimie générale",                            "Pr. Bilongo Anne",       "CM", 2.5),
    ("CHIM", "L1", "CHM112", "Chimie organique I",                         "Dr. Mvondo Eric",        "CM", 2.5),
    ("CHIM", "L1", "CHM113", "Mathématiques pour chimistes",               "Pr. Mbarga Joseph",      "CM", 2.5),
    ("CHIM", "L1", "CHM114", "Physique pour chimistes",                    "Dr. Essomba Robert",     "CM", 2.5),
    ("CHIM", "L1", "CHM115", "TP Chimie générale",                         "Pr. Bilongo Anne",       "TP", 2.5),
    ("CHIM", "L1", "CHM116", "Anglais scientifique I",                     "Mme. Edzimbi Claire",    "TD", 2.5),

    # CHIM L2 — Semestre 3
    ("CHIM", "L2", "CHM211", "Chimie organique II",                        "Dr. Mvondo Eric",        "CM", 2.5),
    ("CHIM", "L2", "CHM212", "Chimie analytique",                          "Pr. Bilongo Anne",       "CM", 2.5),
    ("CHIM", "L2", "CHM213", "Thermodynamique chimique",                   "Dr. Essomba Robert",     "CM", 2.5),
    ("CHIM", "L2", "CHM214", "Chimie inorganique",                         "Dr. Mbida Olivier",      "CM", 2.5),
    ("CHIM", "L2", "CHM215", "Spectroscopie",                              "Dr. Mvondo Eric",        "CM", 2.5),
    ("CHIM", "L2", "CHM216", "TP Chimie analytique",                       "Pr. Bilongo Anne",       "TP", 2.5),

    # CHIM L3 — Semestre 5
    ("CHIM", "L3", "CHM311", "Chimie industrielle",                        "Pr. Bilongo Anne",       "CM", 2.5),
    ("CHIM", "L3", "CHM312", "Génie des procédés",                         "Dr. Mbida Olivier",      "CM", 2.5),
    ("CHIM", "L3", "CHM313", "Cinétique et catalyse",                      "Dr. Mvondo Eric",        "CM", 2.5),
    ("CHIM", "L3", "CHM314", "Chimie de l'environnement",                  "Dr. Essomba Robert",     "CM", 2.5),
    ("CHIM", "L3", "CHM315", "TP Génie des procédés",                      "Dr. Mbida Olivier",      "TP", 2.5),
    ("CHIM", "L3", "CHM316", "Gestion qualité & métrologie",               "Dr. Owona Sylvie",       "TD", 2.5),
]


def construire_fichier(matieres, chemin, titre_feuille):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titre_feuille

    # En-tête
    for col, (header, larg) in enumerate(zip(ENTETES, LARGEURS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color=COULEUR_TEXTE, size=11)
        cell.fill = PatternFill(start_color=COULEUR_EN_TETE, end_color=COULEUR_EN_TETE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = larg
    ws.row_dimensions[1].height = 32

    # Bordure légère
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # Lignes de données
    for r_idx, ligne in enumerate(matieres, start=2):
        for col, val in enumerate(ligne, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.font = Font(color="1A1A1A", size=10)
            cell.alignment = Alignment(horizontal="left" if col in (4, 5) else "center", vertical="center")
            cell.border = border
            if r_idx % 2 == 0:
                cell.fill = PatternFill(start_color=COULEUR_LIGNE_PAIRE, end_color=COULEUR_LIGNE_PAIRE, fill_type="solid")

    ws.freeze_panes = "A2"
    wb.save(chemin)
    print(f"Généré : {chemin} ({len(matieres)} lignes)")


if __name__ == "__main__":
    construire_fichier(MATIERES_TIC,  "/home/neo/Bureau/ChronoFS/import_TIC_S1.xlsx",    "TIC - Semestre 1")
    construire_fichier(MATIERES_CHIM, "/home/neo/Bureau/ChronoFS/import_CHIMIE_S1.xlsx", "Chimie - Semestre 1")
