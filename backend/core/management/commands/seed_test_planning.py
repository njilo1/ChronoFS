"""
Commande de seed pour la simulation de soutenance.

Crée (de façon idempotente) le référentiel — départements, classes (Filiere),
UE et enseignants — extrait des plannings PDF officiels (semaine du 01 au
06/06/2026), puis génère 5 fichiers Excel au format chef de département
(structurellement identiques au template officiel) dans un dossier de sortie.

Objectif : disposer de fichiers de test importables sans erreur pour
démontrer le workflow « import chef → génération DAR » devant le jury.

  python manage.py seed_test_planning
  python manage.py seed_test_planning --out /chemin/vers/test

Aucune donnée n'est supprimée : tout passe par get_or_create. Les effectifs
absents du PDF sont fixés à des valeurs réalistes par niveau.
"""

from __future__ import annotations

import os
import re

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from core.constants import Grade, Niveau, Role, StatutEnseignant, TypeCours, Ville
from core.models import Departement, Enseignant, Filiere, UE, User
from core.services.excel_service import (
    HORAIRES_AFFICHES_HUM,
    JOURS_AFFICHES,
    NIVEAU_CODE_VERS_AFFICHE,
    TYPE_COURS_CODE_VERS_AFFICHE,
    generate_template_excel,
    parse_import_excel,
)

# Effectifs réalistes par défaut (le PDF ne les donne pas).
EFFECTIF_PAR_NIVEAU = {
    Niveau.L1: 80, Niveau.L2: 50, Niveau.L3: 40, Niveau.M1: 20, Niveau.M2: 12,
}

DOSSIER_SORTIE_DEFAUT = '/home/neo/Bureau/ChronoFS/test'

# Chef de département à créer SI le département n'en a pas déjà un.
# (username, mot de passe, nom, grade). On ne remplace jamais un chef existant.
CHEFS = {
    'TIC':  ('chef_tic',  'tic123',  'KENGNI',        Grade.DR),
    'PA':   ('chef_pa',   'pa123',   'NDJOMO',        Grade.DR),
    'CA':   ('chef_ca',   'ca123',   'MBAZE',         Grade.PR),
    'SBM':  ('chef_sbm',  'sbm123',  'EVINA',         Grade.DR),
    'ROSE': ('chef_rose', 'rose123', 'MENGUE MENGUE', Grade.PR),
}

# Raccourcis de créneaux/jours pour des données lisibles.
LUN, MAR, MER, JEU, VEN, SAM = 0, 1, 2, 3, 4, 5
C0, C1, C2, C3 = 0, 1, 2, 3
CM, TD, TP, SEM, PROJ = (
    TypeCours.CM, TypeCours.TD, TypeCours.TP, TypeCours.SEMINAIRE, TypeCours.PROJET,
)
DR, PR, M, MME, ING = Grade.DR, Grade.PR, Grade.M, Grade.MME, Grade.ING
PERM, VAC = StatutEnseignant.PERMANENT, StatutEnseignant.VACATAIRE


# ═══════════════════════════════════════════════════════════════════════════
# DONNÉES — extraites des plannings PDF (campus Ébolowa)
# Structure : département → liste de classes ; chaque classe a ses UE ;
# chaque UE porte son enseignant (nom, grade, statut), son type, et la liste
# des créneaux occupés (jour, créneau).
# ═══════════════════════════════════════════════════════════════════════════
DONNEES = {
    'TIC': {
        'nom': "Technologies de l'Information et de la Communication",
        'classes': [
            ('TIC L1', Niveau.L1, [
                ('TIC142', 'Analyse I', ('ESSONO', DR, PERM), CM,
                 [(LUN, C0), (LUN, C1), (LUN, C2), (MAR, C0), (MAR, C1)]),
                ('TIC122', 'Introduction aux réseaux', ('ASSOUMOU', ING, PERM), CM,
                 [(SAM, C0), (SAM, C1), (SAM, C2), (SAM, C3)]),
            ]),
            ('TIC L2', Niveau.L2, [
                ('TIC224', 'Développement logiciel', ('HATMANN', M, PERM), CM,
                 [(JEU, C1), (JEU, C2), (JEU, C3),
                  (VEN, C0), (VEN, C1), (VEN, C2), (VEN, C3),
                  (SAM, C0), (SAM, C1), (SAM, C2), (SAM, C3)]),
            ]),
            ('TIC L3', Niveau.L3, [
                ('TIC356', "Projet d'intégration", ('KENGNI', DR, PERM), PROJ,
                 [(MAR, C0), (MER, C0), (JEU, C0), (MER, C1), (VEN, C1), (MAR, C2)]),
                ('TIC346', "Développement d'applications mobiles", ('DAOUDA', M, PERM), CM,
                 [(VEN, C0), (SAM, C0), (MAR, C1), (SAM, C1), (LUN, C2), (SAM, C2), (SAM, C3)]),
                ('TIC316', "Management des systèmes d'information", ('ASSOUMOU', ING, PERM), CM,
                 [(LUN, C3), (MAR, C3), (MER, C3), (JEU, C3), (VEN, C3)]),
            ]),
            ('M1 IA et BIG DATA', Niveau.M1, [
                ('TIC458', "Projet d'intégration", ('KENGNI', DR, PERM), PROJ,
                 [(LUN, C0), (LUN, C1), (LUN, C2), (LUN, C3), (SAM, C0), (SAM, C1)]),
            ]),
            ('M2 IA et BIG DATA', Niveau.M2, [
                ('TIC549', 'Apprentissage profond par renforcement', ('DAYANG', PR, PERM), CM,
                 [(LUN, C0), (LUN, C1), (LUN, C2), (LUN, C3),
                  (JEU, C0), (JEU, C1), (JEU, C2), (JEU, C3),
                  (SAM, C0), (SAM, C1), (SAM, C2), (SAM, C3)]),
                ('TIC529', 'IA pour les systèmes Edge et IoT', ('MESSI', DR, PERM), CM,
                 [(MAR, C0), (MAR, C1)]),
                ('TIC559', 'IA fédérée et vie privée', ('FENDJI', PR, PERM), CM,
                 [(MER, C0), (MER, C1), (VEN, C0), (VEN, C1)]),
            ]),
        ],
    },
    'PA': {
        'nom': 'Physique Appliquée',
        'classes': [
            ('Physique Appliquée L1', Niveau.L1, [
                ('PHY142', 'Electrocinétique et magnétostatique', ('NGOUMOU', DR, PERM), CM,
                 [(LUN, C1), (LUN, C3)]),
                ('PHY144', 'Mathématiques pour les physiques 2', ('ESSONO', DR, PERM), CM,
                 [(MER, C0), (MER, C1), (MER, C2), (JEU, C0), (JEU, C1), (JEU, C2)]),
            ]),
            ('PA L2', Niveau.L2, [
                ("PHY242", "Conversion d'Energies", ('NKOLO NKOLO', DR, PERM), CM,
                 [(JEU, C3), (SAM, C2), (SAM, C3)]),
            ]),
            ('Physique Appliquée M2', Niveau.M2, [
                ('PHY542', 'Travaux dirigés', ('ATANGANA', PR, PERM), TD,
                 [(LUN, C0), (LUN, C1), (LUN, C3), (MAR, C2), (SAM, C2), (SAM, C3)]),
            ]),
            # ── Filière de PA basée à MONATÉLÉ (seule filière PA hors Ébolowa) ──
            ('Machinisme Agricole L1', Niveau.L1, Ville.MONATELE, [
                ('MAC142', 'Machine électrique', ('THEPI', DR, PERM), CM,
                 [(j, c) for j in (LUN, MAR, MER, JEU, VEN) for c in (C0, C1, C2, C3)]),
            ]),
            ('Machinisme Agricole L2 & L3', Niveau.L2, Ville.MONATELE, [
                ('MAC242', 'Suivi des réalisations des projets agricoles', ('TOUPOU', ING, PERM), PROJ,
                 [(j, c) for j in (LUN, MAR, MER, JEU, VEN) for c in (C0, C1, C2, C3)]),
            ]),
        ],
    },
    'CA': {
        'nom': 'CHIMIE APPLIQUEE',
        'classes': [
            ('Chimie Appliquée L3', Niveau.L3, [
                ('CHM214', 'Principales fonctions de la chimie organique', ('MBAZE', PR, PERM), CM,
                 [(JEU, C0), (JEU, C1), (JEU, C2), (JEU, C3)]),
                ('CHM316', 'Spectroscopie', ('MBAZE', PR, PERM), CM,
                 [(LUN, C2), (LUN, C3)]),
                ('CHM224', 'Eléments de radiocristallographie', ('NGUIAMBA', DR, PERM), CM,
                 [(LUN, C0), (LUN, C1)]),
                ('CHM376', 'Chimie des eaux', ('GATCHA', DR, PERM), CM,
                 [(MER, C2), (MER, C3)]),
                ('CHM1122', 'Cinétique Chimique', ('MBOUOMBOUO', PR, PERM), CM,
                 [(MAR, C0), (MAR, C1)]),
            ]),
            ('MASTER 1 Chimie Appliquée', Niveau.M1, [
                ('CHM421', 'Chimie de coordination et Modélisation moléculaire', ('NGUIAMBA', DR, PERM), CM,
                 [(MER, C0), (MER, C1)]),
                ('CHF316', 'Formulation et pharmacologie', ('MVONDO', PR, PERM), CM,
                 [(MAR, C2), (MAR, C3)]),
                ('CHI316', 'Introduction au génie des procédés industriels', ('NGUIAMBA', DR, PERM), CM,
                 [(JEU, C0), (JEU, C1)]),
                ('CHF411', "Pharmacognosie et molécules d'intérêt thérapeutique", ('MVONDO', PR, PERM), CM,
                 [(VEN, C0), (VEN, C1)]),
            ]),
        ],
    },
    'SBM': {
        'nom': 'science biomedical',
        'classes': [
            # Tronc commun L1 (réutilise la filière existante "science biomedical")
            ('science biomedical', Niveau.L1, [
                ('SBM112', 'Anatomie II et Physiologie II', ('EVINA', DR, PERM), CM,
                 [(LUN, C0), (LUN, C1), (LUN, C2), (LUN, C3)]),
                ('SBM152', 'Chimie Générale et Biophysique', ('GATCHA', DR, PERM), CM,
                 [(MAR, C0), (MAR, C1), (MAR, C2), (MAR, C3)]),
                ('SBM336', 'Bactériologie 2', ('ATEBA', DR, PERM), CM,
                 [(JEU, C0), (JEU, C1)]),
            ]),
            ('RIM L2', Niveau.L2, [
                ('RIM316', 'Radio pédiatrie', ('EDZIMBI', DR, PERM), CM,
                 [(MAR, C0), (MAR, C1)]),
                ('RIM356', 'Epidémiologie et Economie de la Santé', ('NKE', MME, PERM), CM,
                 [(LUN, C0)]),
                ('RIM412', 'Protocole TDM et Technique échographique', ('EDZIMBI', DR, PERM), CM,
                 [(LUN, C2), (LUN, C3)]),
                ('RIM422', 'Informatique médicale', ('KENGNI', DR, PERM), CM,
                 [(JEU, C1), (JEU, C2), (JEU, C3)]),
            ]),
            ('Biotechnologie et Pharmacognosie L2', Niveau.L2, [
                ('BIP2441', 'Fonction de Nutrition', ('MEZUI', PR, PERM), CM,
                 [(MER, C0), (MER, C1)]),
                ('BIP264', 'Aspects légaux, Ethique et Sécurité des Biotechnologies', ('NGUELE', DR, PERM), CM,
                 [(SAM, C0), (SAM, C1)]),
            ]),
            ('M1 Pharmacognosie', Niveau.M1, [
                ('PBS4281', 'Maladies bactériennes', ('MEZUI', PR, PERM), CM,
                 [(LUN, C0), (LUN, C1)]),
                ('PBS4482', "Procédure pour l'autorisation de mise sur le marché", ('NKOO', DR, VAC), CM,
                 [(JEU, C0), (JEU, C1)]),
                ('PBS4581', 'Assurance et contrôle qualité des médicaments naturels', ('MANG', DR, PERM), CM,
                 [(VEN, C0), (VEN, C1)]),
            ]),
        ],
    },
    'ROSE': {
        'nom': 'Recherche Opérationnelle, Statistique et Économétrie',
        'classes': [
            ('ROSE L2', Niveau.L2, [
                ('MAT214', 'Algèbre multilinéaire', ('MENGUE MENGUE', PR, PERM), CM,
                 [(JEU, C0), (JEU, C1), (JEU, C2), (JEU, C3),
                  (VEN, C0), (VEN, C1), (VEN, C2), (VEN, C3)]),
            ]),
            ('ROSE L3', Niveau.L3, [
                ('MAT336', 'Variables Complexes', ('TCHINDA', DR, PERM), CM,
                 [(LUN, C0), (MAR, C0), (MER, C0), (MER, C1)]),
                ('MAT366', 'Introduction à la Recherche Opérationnelle', ('ETOA ETOA', PR, VAC), CM,
                 [(MAR, C1)]),
                ('MAT356', 'Calcul différentiel', ('ESSONO', DR, PERM), CM,
                 [(MAR, C2), (MAR, C3)]),
            ]),
            ('Master 1 ROSE', Niveau.M1, [
                ('MAT457', "Outils mathématiques pour l'actuariat", ('DJANSI', DR, VAC), CM,
                 [(JEU, C0), (JEU, C1), (JEU, C2), (JEU, C3),
                  (VEN, C0), (VEN, C1), (VEN, C2), (VEN, C3),
                  (SAM, C0), (SAM, C1), (SAM, C2), (SAM, C3)]),
                ('MAT451', 'Systèmes dynamiques', ('TCHINDA', DR, PERM), CM,
                 [(LUN, C1)]),
            ]),
        ],
    },
}


def _classes(info):
    """Itère les classes d'un département en normalisant la ville.

    Une classe est (nom, niveau, ues) — ville Ébolowa par défaut — ou
    (nom, niveau, ville, ues) pour forcer un autre campus (ex. Monatélé).
    """
    for entry in info['classes']:
        if len(entry) == 4:
            nom, niveau, ville, ues = entry
        else:
            nom, niveau, ues = entry
            ville = Ville.EBOLOWA
        yield nom, niveau, ville, ues


class Command(BaseCommand):
    help = "Seed du référentiel + génération des 5 Excel de test (semaine 01-06/06/2026)."

    def add_arguments(self, parser):
        parser.add_argument('--out', default=DOSSIER_SORTIE_DEFAUT,
                            help='Dossier de sortie des fichiers Excel.')

    def handle(self, *args, **opts):
        out = opts['out']
        os.makedirs(out, exist_ok=True)

        for code, info in DONNEES.items():
            dept = self._seed_departement(code, info)
            self._seed_chef(code, dept)
            chemin, rapport = self._generer_excel(dept, info, out)
            statut = '✅ OK' if rapport.ok and rapport.lignes_erreur == 0 else f'⚠️ {rapport.lignes_erreur} erreur(s)'
            self.stdout.write(
                f'{code:5} → {os.path.basename(chemin):28} '
                f'| {rapport.lignes_ok} lignes valides | {statut}'
            )
            for err in rapport.erreurs[:5]:
                self.stdout.write(self.style.WARNING(f'        ligne {err["ligne"]}: {err["message"]}'))

        self.stdout.write(self.style.SUCCESS(f'\nTerminé. Fichiers dans : {out}'))

    # ── Seed référentiel ─────────────────────────────────────────────────────
    def _seed_departement(self, code, info):
        dept, _ = Departement.objects.get_or_create(code=code, defaults={'nom': info['nom']})

        for classe_nom, niveau, ville, ues in _classes(info):
            # Code unique par classe (slug) — la contrainte d'unicité porte sur
            # (code, niveau, ville). On évite ainsi toute collision avec une
            # filière existante du même département au même niveau.
            code_filiere = re.sub(r'[^A-Za-z0-9]', '', classe_nom).upper()[:20]
            filiere, _ = Filiere.objects.get_or_create(
                nom=classe_nom, niveau=niveau,
                defaults={
                    'code': code_filiere, 'departement': dept, 'ville': ville,
                    'effectif': EFFECTIF_PAR_NIVEAU[niveau],
                },
            )
            for ue_code, intitule, (ens_nom, grade, statut), _type, _slots in ues:
                # UE : code unique global → get_or_create par code seul.
                UE.objects.get_or_create(
                    code=ue_code,
                    defaults={'intitule': intitule, 'filiere': filiere},
                )
                ens, _ = Enseignant.objects.get_or_create(
                    nom=ens_nom, grade=grade, defaults={'statut': statut},
                )
                ens.departements.add(dept)
                # Matricule officiel pour les permanents (dérivé de l'id → unique
                # et stable entre deux exécutions). Les vacataires n'en ont pas.
                if ens.statut == StatutEnseignant.PERMANENT and not ens.matricule:
                    lettre = chr(ord('A') + ens.id % 26)
                    ens.matricule = f'{ens.id:07d}{lettre}'
                    ens.save(update_fields=['matricule'])
        return dept

    # ── Chef de département ───────────────────────────────────────────────────
    def _seed_chef(self, code, dept):
        """Garantit qu'un département a un chef. N'écrase jamais un chef existant."""
        if dept.chefs.filter(role=Role.CHEF_DEPT).exists():
            return  # déjà un chef → on ne touche à rien
        username, mdp, nom, grade = CHEFS[code]
        chef, cree = User.objects.get_or_create(
            username=username,
            defaults={
                'role': Role.CHEF_DEPT,
                'departement': dept,
                'last_name': nom,
                'grade': grade,
                'is_staff': True,
            },
        )
        if cree:
            chef.set_password(mdp)
            chef.save()
            self.stdout.write(self.style.SUCCESS(
                f'        chef créé pour {code} : {username} / {mdp}'
            ))

    # ── Génération + validation d'un Excel ───────────────────────────────────
    def _generer_excel(self, dept, info, out):
        contenu, _nom = generate_template_excel(dept)
        wb = load_workbook(filename=__import__('io').BytesIO(contenu))
        ws = wb['Planning']

        rows = []
        for classe_nom, niveau, _ville, ues in _classes(info):
            for ue_code, intitule, (ens_nom, grade, _statut), _type, slots in ues:
                label_ens = f'{Grade(grade).label} {ens_nom}'
                for (jour, creneau) in slots:
                    rows.append((niveau, classe_nom, ue_code, intitule, label_ens,
                                 jour, creneau, _type))

        # Écriture à partir de la ligne 2 (remplace la ligne d'exemple).
        for i, (niveau, classe, ue_code, intitule, ens, jour, creneau, type_c) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=NIVEAU_CODE_VERS_AFFICHE[niveau])
            ws.cell(row=i, column=2, value=classe)
            ws.cell(row=i, column=3, value=ue_code)
            ws.cell(row=i, column=4, value=intitule)
            ws.cell(row=i, column=5, value=ens)
            ws.cell(row=i, column=6, value=JOURS_AFFICHES[jour])
            ws.cell(row=i, column=7, value=HORAIRES_AFFICHES_HUM[creneau])
            ws.cell(row=i, column=8, value=TYPE_COURS_CODE_VERS_AFFICHE[type_c])
            ws.cell(row=i, column=9, value='')

        nom_fichier = f'Planning_{dept.code}_S23_2026.xlsx'
        chemin = os.path.join(out, nom_fichier)
        wb.save(chemin)

        # Validation : on relit le fichier produit avec le vrai parser.
        with open(chemin, 'rb') as f:
            rapport = parse_import_excel(f.read(), dept)
        return chemin, rapport
