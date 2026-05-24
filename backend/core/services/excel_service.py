"""
Service Excel — génération du template et parsing des imports chef.

Le template produit ICI doit être structurellement identique à celui de
référence : `docs/references/Planning_TIC_S21_2026_template.xlsx`.

3 onglets :
- Planning      : visible, saisie utilisateur (9 colonnes, 7 validations)
- Instructions  : visible, guide pas-à-pas
- Référentiel   : MASQUÉ, source des listes déroulantes + VLOOKUP

Le téléchargement est dynamique : chaque GET regénère depuis la BDD,
pour que toute UE/enseignant fraîchement créé par le chef apparaisse
immédiatement.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from core.constants import (
    Creneau,
    HORAIRES_CRENEAUX,
    Jour,
    Niveau,
    TypeCours,
)
from core.models import (
    Departement,
    Enseignant,
    Filiere,
    Semaine,
    UE,
)


# ── Palette FSChrono v2 (bleu UEB) ───────────────────────────────────────────
# Préfixe 'FF' = alpha 100 (opaque). Openpyxl conserve l'alpha tel quel à
# la relecture ; sans préfixe il met '00' (transparent) ce qui rend les
# comparaisons d'égalité avec le fichier de référence fausses.
COULEUR_BLEU_UEB    = 'FF1E3A8A'  # en-tête tableau, sections titres
COULEUR_BLEU_CLAIR  = 'FFDBEAFE'
COULEUR_TEXTE_BLANC = 'FFFFFFFF'
COULEUR_GRIS_FOND   = 'FFF3F4F6'  # ligne exemple "à supprimer"
COULEUR_TEXTE_GRIS  = 'FF6B7280'  # texte secondaire
COULEUR_TEXTE_FONCE = 'FF1F2937'  # texte principal
COULEUR_JAUNE_AUTO  = 'FFFEF3C7'  # cellules auto-remplies (VLOOKUP)


# ── Libellés affichés dans le template ───────────────────────────────────────
# Note : on garde "LICENCE 1" en majuscules comme le template officiel.
NIVEAUX_AFFICHES: list[str] = [
    'LICENCE 1', 'LICENCE 2', 'LICENCE 3', 'MASTER 1', 'MASTER 2',
]
# Conversion entre libellé affiché et code interne (L1, L2, ...)
NIVEAU_AFFICHE_VERS_CODE: dict[str, str] = {
    'LICENCE 1': Niveau.L1,
    'LICENCE 2': Niveau.L2,
    'LICENCE 3': Niveau.L3,
    'MASTER 1':  Niveau.M1,
    'MASTER 2':  Niveau.M2,
}
NIVEAU_CODE_VERS_AFFICHE: dict[str, str] = {v: k for k, v in NIVEAU_AFFICHE_VERS_CODE.items()}

JOURS_AFFICHES = [j.label for j in Jour]                       # Lundi..Samedi
JOUR_AFFICHE_VERS_INDEX: dict[str, int] = {j.label: j.value for j in Jour}

HORAIRES_AFFICHES = [f'{deb}-{fin}' for deb, fin in HORAIRES_CRENEAUX.values()]
HORAIRE_AFFICHE_VERS_INDEX: dict[str, int] = {
    f'{deb}-{fin}': c for c, (deb, fin) in HORAIRES_CRENEAUX.items()
}
# Le template affiche les horaires "7h30-10h00" et pas "07:30-10:00"
def _hum(horaire: str) -> str:
    """ '07:30-10:00' → '7h30-10h00' """
    return horaire.replace('07:', '7h').replace('08:', '8h').replace('09:', '9h') \
                  .replace('10:', '10h').replace('12:', '12h').replace('13:', '13h') \
                  .replace('15:', '15h').replace('18:', '18h')

HORAIRES_AFFICHES_HUM = [_hum(h) for h in HORAIRES_AFFICHES]
HORAIRE_HUM_VERS_INDEX: dict[str, int] = {
    _hum(h): idx for h, idx in HORAIRE_AFFICHE_VERS_INDEX.items()
}

TYPES_COURS_AFFICHES: dict[str, str] = {
    'CM':        TypeCours.CM,
    'TD':        TypeCours.TD,
    'TP':        TypeCours.TP,
    'Séminaire': TypeCours.SEMINAIRE,
    'Projet':    TypeCours.PROJET,
}
TYPE_COURS_CODE_VERS_AFFICHE = {v: k for k, v in TYPES_COURS_AFFICHES.items()}

# Marqueur Excel pour qu'on saute la ligne d'exemple à l'import
MARQUEUR_EXEMPLE = "Ligne d'exemple - À SUPPRIMER"


# ═════════════════════════════════════════════════════════════════════════════
# GÉNÉRATION DU TEMPLATE
# ═════════════════════════════════════════════════════════════════════════════

def generate_template_excel(departement: Departement, semaine: Optional[Semaine] = None) -> tuple[bytes, str]:
    """
    Génère le template Excel pour ce département et cette semaine.

    Retourne (contenu_bytes, nom_de_fichier_suggéré).
    """
    filieres = list(
        Filiere.objects.filter(departement=departement).order_by('niveau', 'ville', 'code')
    )
    ues = list(
        UE.objects.filter(filiere__departement=departement)
        .select_related('filiere')
        .order_by('code')
    )
    enseignants = list(
        Enseignant.objects.filter(departements=departement)
        .order_by('nom', 'grade')
    )

    # Détecter les couples (code, niveau) présents dans plusieurs villes :
    # pour ceux-là on affichera "(Ébolowa)" / "(Monatélé)" à la fin du
    # nom de classe.
    occurrences: dict[tuple[str, str], int] = {}
    for f in filieres:
        occurrences[(f.code, f.niveau)] = occurrences.get((f.code, f.niveau), 0) + 1

    def libelle_classe(f: Filiere) -> str:
        # On utilise le `nom` pédagogique de la filière (ex. "M1 IA et BIG DATA").
        if occurrences[(f.code, f.niveau)] > 1:
            return f'{f.nom} ({f.get_ville_display()})'
        return f.nom

    wb = Workbook()

    # ── Onglet Planning ──────────────────────────────────────────────────────
    ws_planning: Worksheet = wb.active
    ws_planning.title = 'Planning'
    _construire_onglet_planning(ws_planning, ues)

    # ── Onglet Instructions ──────────────────────────────────────────────────
    ws_inst = wb.create_sheet('Instructions')
    _construire_onglet_instructions(ws_inst, departement, semaine)

    # ── Onglet Référentiel (masqué) ──────────────────────────────────────────
    ws_ref = wb.create_sheet('Référentiel')
    _construire_onglet_referentiel(ws_ref, filieres, ues, enseignants, libelle_classe)
    ws_ref.sheet_state = 'hidden'

    # ── Validations data après création du Référentiel (références correctes) ─
    _appliquer_validations(ws_planning, filieres, ues, enseignants)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nom_fichier = _nom_fichier_template(departement, semaine)
    return buf.getvalue(), nom_fichier


# ── Sous-fonctions de construction ───────────────────────────────────────────
def _nom_fichier_template(dept: Departement, semaine: Optional[Semaine]) -> str:
    if semaine:
        num_iso = semaine.date_debut.isocalendar()[1]
        return f'Planning_{dept.code}_S{num_iso:02d}_{semaine.date_debut.year}.xlsx'
    return f'Planning_{dept.code}.xlsx'


def _construire_onglet_planning(ws: Worksheet, ues: list[UE]):
    """En-tête bleu UEB + ligne exemple grisée + colonne D auto (VLOOKUP)."""
    entetes = [
        ('Niveau',           14),
        ('Classe',           20),
        ('Code UE',          12),
        ('Cours (intitulé)', 45),
        ('Enseignant',       22),
        ('Jour',             12),
        ('Horaire',          14),
        ('Type Cours',       12),
        ('Observations',     32),
    ]

    fill_bleu = PatternFill(start_color=COULEUR_BLEU_UEB, end_color=COULEUR_BLEU_UEB, fill_type='solid')
    font_blanc = Font(bold=True, color=COULEUR_TEXTE_BLANC, size=11)
    align_centre = Alignment(horizontal='center', vertical='center')
    bordure_fine = Border(*([Side(style='thin', color='D1D5DB')] * 4))

    # Ligne 1 : entêtes
    for col, (titre, largeur) in enumerate(entetes, 1):
        cell = ws.cell(row=1, column=col, value=titre)
        cell.fill      = fill_bleu
        cell.font      = font_blanc
        cell.alignment = align_centre
        cell.border    = bordure_fine
        ws.column_dimensions[get_column_letter(col)].width = largeur
    ws.row_dimensions[1].height = 24

    # Ligne 2 : exemple "à supprimer", style discret
    fill_gris = PatternFill(start_color=COULEUR_GRIS_FOND, end_color=COULEUR_GRIS_FOND, fill_type='solid')
    font_exemple = Font(italic=True, color=COULEUR_TEXTE_GRIS, size=10)
    exemple = [
        'LICENCE 3', 'TIC L3', 'TIC336', 'Re-ingénierie des processus métiers',
        'M DAOUDA', 'Lundi', '7h30-10h00', 'CM', MARQUEUR_EXEMPLE,
    ]
    for col, val in enumerate(exemple, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = fill_gris
        cell.font = font_exemple
        cell.border = bordure_fine

    # Lignes 3..102 : colonne D auto-remplie par VLOOKUP sur la BDD UE
    fin_ref = 1 + len(ues) if ues else 13  # protection si dept vide
    formule_lookup_tpl = f'=IFERROR(VLOOKUP(C{{row}},Référentiel!$C$2:$D${fin_ref},2,FALSE()),"")'
    fill_auto = PatternFill(start_color=COULEUR_JAUNE_AUTO, end_color=COULEUR_JAUNE_AUTO, fill_type='solid')
    protected = Protection(locked=True)
    for r in range(3, 103):
        cell = ws.cell(row=r, column=4, value=formule_lookup_tpl.format(row=r))
        cell.fill       = fill_auto
        cell.protection = protected
        cell.font       = Font(size=10, color=COULEUR_TEXTE_FONCE)

    # Volets figés à partir de la ligne 3 (entête + exemple toujours visibles)
    ws.freeze_panes = 'A3'


def _construire_onglet_instructions(ws: Worksheet, dept: Departement, semaine: Optional[Semaine]):
    """Guide pas-à-pas pour le chef de département."""
    ws.column_dimensions['A'].width = 100

    def _ecrire(row: int, texte: str, *, bold=False, size=11, color=COULEUR_TEXTE_FONCE):
        cell = ws.cell(row=row, column=1, value=texte)
        cell.font = Font(bold=bold, size=size, color=color)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    semestre_txt = f'Semestre : {semaine.semestre}' if semaine else 'Semestre : —'
    annee_txt    = (
        f'Année académique : {semaine.annee_academique.libelle}'
        if semaine and semaine.annee_academique_id
        else 'Année académique : —'
    )
    semaine_txt = (
        f'Semaine : du {semaine.date_debut:%d/%m/%Y} au {semaine.date_fin:%d/%m/%Y}'
        if semaine else 'Semaine : à venir'
    )

    lignes: list[tuple[str, dict]] = [
        ('FSChrono — Guide de remplissage du planning hebdomadaire', {'bold': True, 'size': 18}),
        ('', {}),
        (f'Département : {dept.code} ({dept.nom})', {'bold': True, 'size': 12}),
        (f'{annee_txt}   |   {semestre_txt}',       {'bold': True, 'size': 12}),
        (semaine_txt,                                {'bold': True, 'size': 12, 'color': COULEUR_TEXTE_GRIS}),
        ('', {}),
        ('📋 PROCÉDURE', {'bold': True, 'size': 13, 'color': COULEUR_BLEU_UEB}),
        ("1. Ouvrez l'onglet 'Planning' (onglets en bas à gauche).", {}),
        ('2. SUPPRIMEZ la ligne 2 (exemple grisé) avant de commencer.', {}),
        ('3. Remplissez une ligne par cours à programmer cette semaine.', {}),
        ('4. Utilisez les listes déroulantes pour éviter les fautes de frappe.', {}),
        ("5. La colonne 'Cours (intitulé)' se remplit AUTOMATIQUEMENT quand vous choisissez le Code UE.", {}),
        ("6. Si un cours dure 2 créneaux, créez 2 lignes (ex. 7h30-10h00 ET 10h15-12h45).", {}),
        ("7. Enregistrez le fichier et envoyez-le via l'application FSChrono.", {}),
        ('', {}),
        ('📊 EXPLICATION DES COLONNES', {'bold': True, 'size': 13, 'color': COULEUR_BLEU_UEB}),
        ('', {}),
        ('• Niveau (obligatoire)        : LICENCE 1, LICENCE 2, LICENCE 3, MASTER 1, MASTER 2', {}),
        ('• Classe (obligatoire)        : nom de la classe pédagogique (ex. TIC L3, M1 IA et BIG DATA).', {}),
        ('                                Pour les filières présentes dans 2 villes, la ville est ajoutée.', {}),
        ('• Code UE (obligatoire)       : code de l\'unité d\'enseignement (ex. TIC336)', {}),
        ('• Cours (AUTO)                : intitulé rempli automatiquement, ne pas modifier', {}),
        ("• Enseignant (obligatoire)    : choisissez dans la liste ou 'Non assigné' si pas de prof désigné", {}),
        ('• Jour (obligatoire)          : Lundi à Samedi UNIQUEMENT (PAS de cours le Dimanche)', {}),
        ('• Horaire (obligatoire)       : un des 4 créneaux journaliers', {}),
        ('• Type Cours (obligatoire)    : CM, TD, TP, Séminaire ou Projet', {}),
        ("• Observations (optionnel)    : précisions (ex. 'cours en ligne', 'besoin de labo')", {}),
        ('', {}),
        ('⚠️ RÈGLES IMPORTANTES', {'bold': True, 'size': 13, 'color': COULEUR_BLEU_UEB}),
        ('', {}),
        ("• Une classe ne peut pas avoir 2 cours au même créneau — le système refuse.", {}),
        ("• Un enseignant ne peut pas être à 2 endroits en même temps.", {}),
        ("• Un cours absent de votre fichier n'apparaîtra PAS dans le planning final de la semaine.", {}),
        ("• Vous n'attribuez PAS de salle, c'est le rôle de la DAR.", {}),
        ("• L'algorithme choisit la salle adaptée selon l'effectif, le type de cours et la ville.", {}),
        ("• Pour les profs inter-villes : au moins 1 créneau d'écart entre Ébolowa et Monatélé.", {}),
        ('', {}),
        ('🔧 EN CAS DE PROBLÈME', {'bold': True, 'size': 13, 'color': COULEUR_BLEU_UEB}),
        ('', {}),
        ("• Code UE manquant ? Vous pouvez l'ajouter dans 'Mon département > UE' avant le téléchargement.", {}),
        ("• Nouvel enseignant ? 'Mon département > Enseignants' permet de l'ajouter, puis re-téléchargez.", {}),
        ("• Question technique ? Contactez la DAR via les coordonnées habituelles.", {}),
        ('', {}),
        ('───────────────────────────────────────────────────────', {'color': COULEUR_TEXTE_GRIS}),
        ("Université d'Ébolowa — Faculté des Sciences — DAARS", {'color': COULEUR_TEXTE_GRIS}),
        ('FSChrono v2.0', {'color': COULEUR_TEXTE_GRIS}),
    ]
    for i, (texte, kwargs) in enumerate(lignes, start=1):
        _ecrire(i, texte, **kwargs)


def _construire_onglet_referentiel(
    ws: Worksheet,
    filieres: list[Filiere],
    ues: list[UE],
    enseignants: list[Enseignant],
    libelle_classe,
):
    """
    Source des listes déroulantes ET de la VLOOKUP. Masqué dans l'UI mais
    présent dans le fichier — Excel a besoin que la feuille existe.
    """
    en_tetes = ['Niveaux', 'Classes', 'Codes UE', 'Intitulés UE',
                'Enseignants', 'Jours', 'Horaires', 'Types Cours']
    largeurs = [18, 28, 12, 45, 22, 12, 14, 12]
    fill_bleu = PatternFill(start_color=COULEUR_BLEU_UEB, end_color=COULEUR_BLEU_UEB, fill_type='solid')
    font_blanc = Font(bold=True, color=COULEUR_TEXTE_BLANC, size=11)
    for col, (titre, larg) in enumerate(zip(en_tetes, largeurs), 1):
        cell = ws.cell(row=1, column=col, value=titre)
        cell.fill = fill_bleu
        cell.font = font_blanc
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = larg

    # Colonne A : Niveaux (LICENCE 1..MASTER 2)
    for i, niveau in enumerate(NIVEAUX_AFFICHES, start=2):
        ws.cell(row=i, column=1, value=niveau)

    # Colonne B : Classes
    for i, f in enumerate(filieres, start=2):
        ws.cell(row=i, column=2, value=libelle_classe(f))

    # Colonnes C/D : Codes UE / Intitulés UE — alignés pour le VLOOKUP
    for i, ue in enumerate(ues, start=2):
        ws.cell(row=i, column=3, value=ue.code)
        ws.cell(row=i, column=4, value=ue.intitule)

    # Colonne E : Enseignants (avec 'Non assigné' en tête)
    enseignants_labels = ['Non assigné'] + [f'{e.get_grade_display()} {e.nom}' for e in enseignants]
    for i, lib in enumerate(enseignants_labels, start=2):
        ws.cell(row=i, column=5, value=lib)

    # Colonne F : Jours
    for i, jour in enumerate(JOURS_AFFICHES, start=2):
        ws.cell(row=i, column=6, value=jour)

    # Colonne G : Horaires (format humanisé : 7h30-10h00)
    for i, h in enumerate(HORAIRES_AFFICHES_HUM, start=2):
        ws.cell(row=i, column=7, value=h)

    # Colonne H : Types Cours
    for i, t in enumerate(TYPES_COURS_AFFICHES.keys(), start=2):
        ws.cell(row=i, column=8, value=t)


def _appliquer_validations(ws: Worksheet, filieres, ues, enseignants):
    """7 validations data sur l'onglet Planning, références → onglet Référentiel."""

    def ajouter(col_letter: str, ref_col: str, nb_items: int, libelle: str):
        # On garde au moins 1 item pour ne pas générer une plage vide
        n = max(1, nb_items)
        formule = f"=Référentiel!${ref_col}$2:${ref_col}${1 + n}"
        dv = DataValidation(
            type='list',
            formula1=formule,
            allow_blank=True,
            errorTitle='Valeur invalide',
            error=f"Choisissez {libelle} dans la liste déroulante.",
            promptTitle=libelle,
            prompt=f"Sélectionnez {libelle}.",
        )
        dv.add(f'{col_letter}3:{col_letter}102')
        ws.add_data_validation(dv)

    ajouter('A', 'A', len(NIVEAUX_AFFICHES),                  "un niveau")
    ajouter('B', 'B', len(filieres),                          "une classe")
    ajouter('C', 'C', len(ues),                               "un code UE")
    ajouter('E', 'E', 1 + len(enseignants),                   "un enseignant")  # +1 pour "Non assigné"
    ajouter('F', 'F', len(JOURS_AFFICHES),                    "un jour")
    ajouter('G', 'G', len(HORAIRES_AFFICHES_HUM),             "un horaire")
    ajouter('H', 'H', len(TYPES_COURS_AFFICHES),              "un type de cours")


# ═════════════════════════════════════════════════════════════════════════════
# PARSING D'UN IMPORT EXCEL
# ═════════════════════════════════════════════════════════════════════════════

@dataclass
class LigneParsee:
    """Une ligne de l'onglet Planning, prête à devenir une DemandeCours."""

    ligne_num:        int
    filiere_id:       Optional[int] = None
    ue_id:            Optional[int] = None
    enseignant_id:    Optional[int] = None
    jour:             Optional[int] = None
    creneau:          Optional[int] = None
    type_cours:       Optional[str] = None
    observations:     str = ''
    effectif_declare: int = 0


@dataclass
class RapportParsing:
    """Résultat du parsing d'un fichier Excel — utilisé en preview et import."""

    ok:             bool                  = True
    lignes_ok:      int                   = 0
    lignes_erreur:  int                   = 0
    lignes_vides:   int                   = 0
    erreurs:        list[dict]            = field(default_factory=list)
    lignes_valides: list[LigneParsee]     = field(default_factory=list)

    def ajouter_erreur(self, ligne: int, message: str):
        self.erreurs.append({'ligne': ligne, 'message': message})
        self.lignes_erreur += 1
        self.ok = False

    def to_json(self) -> dict:
        return {
            'ok':             self.ok,
            'lignes_ok':      self.lignes_ok,
            'lignes_erreur':  self.lignes_erreur,
            'lignes_vides':   self.lignes_vides,
            'erreurs':        self.erreurs,
        }


def parse_import_excel(fichier_bytes: bytes, departement: Departement) -> RapportParsing:
    """
    Lit l'onglet 'Planning' d'un fichier .xlsx et renvoie un rapport.

    Validations :
    - Niveau et Classe doivent appartenir au département
    - Code UE doit exister dans le département
    - Enseignant doit exister (ou 'Non assigné' → null)
    - Jour ∈ {Lundi..Samedi} — dimanche refusé explicitement
    - Horaire ∈ {7h30-10h00, 10h15-12h45, 13h00-15h30, 15h45-18h15}
    - Type Cours ∈ {CM, TD, TP, Séminaire, Projet}

    Tous les messages sont en langage naturel — pas de jargon technique.
    Le rapport contient les lignes valides (prêtes à être converties en
    DemandeCours) et les erreurs détaillées.
    """
    import openpyxl

    rapport = RapportParsing()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(fichier_bytes), data_only=True)
    except Exception as e:
        rapport.ajouter_erreur(0, f"Le fichier n'a pas pu être ouvert : {e}.")
        return rapport

    if 'Planning' not in wb.sheetnames:
        rapport.ajouter_erreur(0, "L'onglet 'Planning' est introuvable. Avez-vous bien utilisé le modèle ?")
        return rapport

    ws = wb['Planning']

    # Pré-charger les index lookup pour ce département
    filieres_par_libelle = _index_filieres_par_libelle(departement)
    ues_par_code         = {ue.code.upper(): ue for ue in
                            UE.objects.filter(filiere__departement=departement).select_related('filiere')}
    enseignants_par_lib  = _index_enseignants_par_libelle(departement)

    # Détecter les triplets (classe, jour, créneau) en double dans le fichier
    seen_classe_creneau: set[tuple[int, int, int]] = set()
    # Détecter les triplets (enseignant, jour, créneau) en double dans le fichier
    seen_ens_creneau: set[tuple[int, int, int]] = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Sauter la ligne d'exemple (ligne 2 marquée dans observations)
        observations = str(row[8]).strip() if row[8] else ''
        if observations == MARQUEUR_EXEMPLE:
            continue
        # Ligne complètement vide
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row[:8]):
            rapport.lignes_vides += 1
            continue

        ligne = _parser_une_ligne(
            row_idx, row,
            filieres_par_libelle, ues_par_code, enseignants_par_lib,
            rapport,
        )
        if ligne is None:
            continue

        # Validation intra-fichier : pas de double pour une classe au même créneau
        cle_classe = (ligne.filiere_id, ligne.jour, ligne.creneau)
        if cle_classe in seen_classe_creneau:
            rapport.ajouter_erreur(
                row_idx,
                f"Cette classe a déjà un cours à ce créneau dans votre fichier "
                f"({JOURS_AFFICHES[ligne.jour]} {HORAIRES_AFFICHES_HUM[ligne.creneau]}). "
                "Une classe ne peut suivre qu'un cours à la fois.",
            )
            continue
        seen_classe_creneau.add(cle_classe)

        # Pas de double pour un enseignant au même créneau (sauf "Non assigné")
        if ligne.enseignant_id is not None:
            cle_ens = (ligne.enseignant_id, ligne.jour, ligne.creneau)
            if cle_ens in seen_ens_creneau:
                rapport.ajouter_erreur(
                    row_idx,
                    f"Cet enseignant est déjà inscrit à ce créneau dans votre fichier "
                    f"({JOURS_AFFICHES[ligne.jour]} {HORAIRES_AFFICHES_HUM[ligne.creneau]}). "
                    "Un enseignant ne peut faire qu'un cours à la fois.",
                )
                continue
            seen_ens_creneau.add(cle_ens)

        rapport.lignes_valides.append(ligne)
        rapport.lignes_ok += 1

    return rapport


# ── Helpers de parsing ───────────────────────────────────────────────────────
def _index_filieres_par_libelle(dept: Departement) -> dict[str, Filiere]:
    """Indexe les filières d'un dept par le libellé affiché dans Excel."""
    filieres = list(Filiere.objects.filter(departement=dept))
    occurrences: dict[tuple[str, str], int] = {}
    for f in filieres:
        occurrences[(f.code, f.niveau)] = occurrences.get((f.code, f.niveau), 0) + 1

    index = {}
    for f in filieres:
        # Libellé "simple"
        index[f.nom.upper()] = f
        # Si multi-villes, on prend aussi le libellé "Nom (Ville)"
        if occurrences[(f.code, f.niveau)] > 1:
            index[f'{f.nom} ({f.get_ville_display()})'.upper()] = f
    return index


def _index_enseignants_par_libelle(dept: Departement) -> dict[str, Enseignant]:
    """Indexe les enseignants par le libellé "Grade Nom" affiché dans Excel."""
    index: dict[str, Enseignant] = {}
    for e in Enseignant.objects.filter(departements=dept):
        index[f'{e.get_grade_display()} {e.nom}'.upper()] = e
    return index


def _parser_une_ligne(
    row_idx: int,
    row: tuple,
    filieres_idx: dict[str, Filiere],
    ues_idx: dict[str, UE],
    enseignants_idx: dict[str, Enseignant],
    rapport: RapportParsing,
) -> Optional[LigneParsee]:
    """Valide une ligne du fichier. Renvoie None si bloquant."""

    niveau_aff  = _str(row[0])
    classe_aff  = _str(row[1])
    code_ue     = _str(row[2])
    # colonne D ignorée (intitulé auto VLOOKUP)
    enseignant  = _str(row[4])
    jour_aff    = _str(row[5])
    horaire_aff = _str(row[6])
    type_aff    = _str(row[7])
    obs         = _str(row[8])

    erreurs_initiales = rapport.lignes_erreur

    # ── Niveau ───────────────────────────────────────────────────────────────
    if not niveau_aff:
        rapport.ajouter_erreur(row_idx, "Le niveau est manquant.")
    elif niveau_aff.upper() not in (n.upper() for n in NIVEAUX_AFFICHES):
        rapport.ajouter_erreur(
            row_idx,
            f"Le niveau « {niveau_aff} » n'est pas reconnu. "
            f"Utilisez LICENCE 1, LICENCE 2, LICENCE 3, MASTER 1 ou MASTER 2.",
        )

    # ── Classe ───────────────────────────────────────────────────────────────
    filiere = None
    if not classe_aff:
        rapport.ajouter_erreur(row_idx, "La classe est manquante.")
    else:
        filiere = filieres_idx.get(classe_aff.upper())
        if filiere is None:
            rapport.ajouter_erreur(
                row_idx,
                f"La classe « {classe_aff} » n'existe pas dans votre département. "
                "Vérifiez l'orthographe ou demandez à la DAR de l'ajouter.",
            )

    # Cohérence niveau ↔ classe
    if filiere and niveau_aff:
        niveau_attendu = NIVEAU_CODE_VERS_AFFICHE.get(filiere.niveau)
        if niveau_attendu and niveau_aff.upper() != niveau_attendu.upper():
            rapport.ajouter_erreur(
                row_idx,
                f"Incohérence : la classe « {classe_aff} » est de niveau "
                f"{niveau_attendu}, pas {niveau_aff}.",
            )

    # ── Code UE ──────────────────────────────────────────────────────────────
    ue = None
    if not code_ue:
        rapport.ajouter_erreur(row_idx, "Le code UE est manquant.")
    else:
        ue = ues_idx.get(code_ue.upper())
        if ue is None:
            rapport.ajouter_erreur(
                row_idx,
                f"Le code UE « {code_ue} » n'existe pas dans votre département. "
                "Ajoutez-le depuis « Mon département > UE » avant de réessayer.",
            )
        elif filiere and ue.filiere_id != filiere.id:
            rapport.ajouter_erreur(
                row_idx,
                f"L'UE « {code_ue} » est rattachée à une autre classe que « {classe_aff} ». "
                "Choisissez une UE compatible avec cette classe.",
            )

    # ── Enseignant ───────────────────────────────────────────────────────────
    enseignant_obj = None
    if enseignant and enseignant.lower() != 'non assigné':
        enseignant_obj = enseignants_idx.get(enseignant.upper())
        if enseignant_obj is None:
            rapport.ajouter_erreur(
                row_idx,
                f"L'enseignant « {enseignant} » n'est pas reconnu. "
                "Ajoutez-le depuis « Mon département > Enseignants » avant de réessayer.",
            )

    # ── Jour ─────────────────────────────────────────────────────────────────
    jour_idx = None
    if not jour_aff:
        rapport.ajouter_erreur(row_idx, "Le jour est manquant.")
    elif jour_aff.lower() == 'dimanche':
        rapport.ajouter_erreur(
            row_idx,
            "Le dimanche n'est pas autorisé : les cours ont lieu du lundi au samedi uniquement.",
        )
    else:
        jour_idx = next(
            (idx for nom, idx in JOUR_AFFICHE_VERS_INDEX.items() if nom.lower() == jour_aff.lower()),
            None,
        )
        if jour_idx is None:
            rapport.ajouter_erreur(
                row_idx,
                f"Le jour « {jour_aff} » n'est pas reconnu. "
                "Utilisez Lundi, Mardi, Mercredi, Jeudi, Vendredi ou Samedi.",
            )

    # ── Horaire ──────────────────────────────────────────────────────────────
    creneau_idx = None
    if not horaire_aff:
        rapport.ajouter_erreur(row_idx, "L'horaire est manquant.")
    else:
        # On accepte les deux formats : '7h30-10h00' et '07:30-10:00'.
        h_norm = horaire_aff.replace(' ', '').lower()
        creneau_idx = next(
            (idx for nom, idx in HORAIRE_HUM_VERS_INDEX.items() if nom.replace(' ', '').lower() == h_norm),
            None,
        )
        if creneau_idx is None:
            creneau_idx = next(
                (idx for nom, idx in HORAIRE_AFFICHE_VERS_INDEX.items() if nom.replace(' ', '').lower() == h_norm),
                None,
            )
        if creneau_idx is None:
            rapport.ajouter_erreur(
                row_idx,
                f"L'horaire « {horaire_aff} » n'est pas reconnu. "
                f"Utilisez {', '.join(HORAIRES_AFFICHES_HUM)}.",
            )

    # ── Type de cours ────────────────────────────────────────────────────────
    type_code = None
    if not type_aff:
        rapport.ajouter_erreur(row_idx, "Le type de cours est manquant.")
    else:
        for aff, code in TYPES_COURS_AFFICHES.items():
            if aff.lower() == type_aff.lower():
                type_code = code
                break
        if type_code is None:
            rapport.ajouter_erreur(
                row_idx,
                f"Le type de cours « {type_aff} » n'est pas reconnu. "
                f"Utilisez {', '.join(TYPES_COURS_AFFICHES.keys())}.",
            )

    # Si une erreur bloquante est survenue sur cette ligne → on l'écarte
    if rapport.lignes_erreur > erreurs_initiales:
        return None

    return LigneParsee(
        ligne_num        = row_idx,
        filiere_id       = filiere.id,
        ue_id            = ue.id,
        enseignant_id    = enseignant_obj.id if enseignant_obj else None,
        jour             = jour_idx,
        creneau          = creneau_idx,
        type_cours       = type_code,
        observations     = obs,
        effectif_declare = filiere.effectif,
    )


def _str(value) -> str:
    """Convertit en str strippé, '' si None."""
    if value is None:
        return ''
    return str(value).strip()
