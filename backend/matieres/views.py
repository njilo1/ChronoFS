from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Matiere
from .serializers import MatiereSerializer
from filieres.models import Filiere, Niveau, Departement
from enseignants.models import Enseignant
from planification.models import SessionPlanification, ImportDepartement


class MatiereViewSet(viewsets.ModelViewSet):
    """API CRUD pour les matières."""

    queryset = (
        Matiere.objects
        .select_related(
            'niveau__filiere__departement',
            'enseignant',
        )
        .all()
    )
    serializer_class   = MatiereSerializer
    permission_classes = [AllowAny]

    filterset_fields = ['niveau', 'enseignant', 'type_seance']
    search_fields    = ['code', 'intitule']
    ordering_fields  = ['code', 'intitule', 'type_seance']


# ── Constantes de validation ──────────────────────────────
TYPES_VALIDES   = {'CM', 'TD', 'TP'}
NIVEAUX_VALIDES = {'L1', 'L2', 'L3', 'M1', 'M2'}

# Couleurs du template Excel
COULEUR_EN_TETE = '0D1520'   # fond sombre
COULEUR_TEXTE   = 'C9A450'   # or ChronoFS
COULEUR_EXEMPLE = '1A2A3A'   # ligne exemple


def _matcher_enseignant(nom_brut, departement):
    """Retrouve un enseignant par son nom (case-insensitive, ignore 'Dr.', 'Pr.', accents).

    Si introuvable, le crée automatiquement et le rattache au département.
    Renvoie (enseignant, created_now: bool).
    """
    if not nom_brut:
        return None, False

    # Nettoyer : retirer les titres communs et les espaces multiples
    cleaned = nom_brut.strip()
    for prefix in ('Pr.', 'Pr ', 'Dr.', 'Dr ', 'Mr.', 'Mr ', 'Mme.', 'Mme ', 'M.'):
        if cleaned.lower().startswith(prefix.lower()):
            cleaned = cleaned[len(prefix):].strip()

    if not cleaned:
        return None, False

    # Recherche case-insensitive sur nom + prénom concaténés
    for ens in Enseignant.objects.all():
        candidats = [
            (ens.nom or '').lower().strip(),
            f"{ens.prenom or ''} {ens.nom or ''}".lower().strip(),
            f"{ens.nom or ''} {ens.prenom or ''}".lower().strip(),
        ]
        if cleaned.lower() in candidats or any(cleaned.lower() in c for c in candidats if c):
            return ens, False

    # Auto-création : déduire grade éventuel du préfixe
    grade = ''
    raw_lower = nom_brut.lower().strip()
    if raw_lower.startswith(('pr.', 'pr ')):
        grade = 'Pr'
    elif raw_lower.startswith(('dr.', 'dr ')):
        grade = 'Dr'

    parts = cleaned.split(' ', 1)
    nom = parts[0]
    prenom = parts[1] if len(parts) > 1 else ''

    ens = Enseignant.objects.create(
        nom=nom,
        prenom=prenom,
        grade=grade,
        email=f"inconnu.{nom.lower()}.{prenom.lower() if prenom else 'x'}@chronofs.local",
        est_actif=True,
    )
    if departement:
        ens.departements.add(departement)
    return ens, True


class ImportMatieresView(APIView):
    """
    GET  /api/matieres/import/  → télécharger le modèle Excel
    POST /api/matieres/import/  → importer un fichier Excel envoyé par un chef de département

    Body POST (multipart/form-data) :
      - fichier        : fichier .xlsx
      - departement    : ID du département concerné
      - dry_run        : 'true' pour un aperçu sans enregistrement

    Format Excel attendu (1 ligne par UE, plusieurs filières du même dept possibles) :
      filiere_code | niveau | code_ue | intitule | enseignant | type_seance | volume_horaire
    """
    parser_classes     = [MultiPartParser, FormParser]
    permission_classes = [AllowAny]

    # ── GET : générer et retourner le modèle Excel ────────
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Planning Matières'

        entetes = [
            'filiere_code',
            'niveau (L1/L2/L3/M1/M2)',
            'code_ue',
            'intitule',
            'enseignant',
            'type_seance (CM/TD/TP)',
            'volume_horaire (h)',
        ]
        largeurs = [14, 22, 14, 42, 30, 22, 18]

        for col, (header, larg) in enumerate(zip(entetes, largeurs), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color=COULEUR_TEXTE, size=11)
            cell.fill      = PatternFill(start_color=COULEUR_EN_TETE, end_color=COULEUR_EN_TETE, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[cell.column_letter].width = larg
        ws.row_dimensions[1].height = 22

        exemples = [
            ['TIC',     'L1', 'TIC111', 'Algorithmique',         'Dr. Nyabeye',    'CM', 2.5],
            ['TIC',     'L2', 'TIC224', 'Analyse & Conception SI', 'Pr. Mboumboumbo', 'CM', 2.5],
            ['TIC-MON', 'L1', 'TIC111', 'Algorithmique',         'Dr. Nyabeye',    'CM', 2.5],
        ]
        for r_idx, row in enumerate(exemples, start=2):
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=col, value=val)
                cell.font = Font(italic=True, color='5A7A8A', size=10)
                cell.fill = PatternFill(start_color=COULEUR_EXEMPLE, end_color=COULEUR_EXEMPLE, fill_type='solid')

        ws.freeze_panes = 'A2'

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="template_planning_chronofs.xlsx"'
        return response

    # ── POST : analyser et importer le fichier ────────────
    def post(self, request):
        fichier      = request.FILES.get('fichier')
        dept_id      = request.data.get('departement')
        session_id   = request.data.get('session')
        dry_run      = request.data.get('dry_run', 'false').strip().lower() == 'true'

        if not fichier:
            return Response({'error': 'Aucun fichier fourni.'}, status=400)
        if not dept_id:
            return Response({'error': 'Le département est requis.'}, status=400)
        if not session_id:
            return Response(
                {'error': "La session de planification est requise. Créez une session avant d'importer."},
                status=400,
            )

        try:
            departement = Departement.objects.get(pk=dept_id)
        except Departement.DoesNotExist:
            return Response({'error': f'Département ID={dept_id} introuvable.'}, status=400)

        try:
            session = SessionPlanification.objects.get(pk=session_id)
        except SessionPlanification.DoesNotExist:
            return Response({'error': f'Session ID={session_id} introuvable.'}, status=400)

        if session.etat in ('genere', 'publie', 'archive'):
            return Response(
                {'error': f"Cette session est {session.get_etat_display().lower()}. Plus d'import possible."},
                status=400,
            )

        # Récupérer (ou créer en sécurité) la ligne de suivi pour ce dept dans cette session.
        suivi, _ = ImportDepartement.objects.get_or_create(
            session=session,
            departement=departement,
        )

        try:
            wb = openpyxl.load_workbook(BytesIO(fichier.read()), data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Fichier Excel invalide : {e}'}, status=400)

        # Charger les filières + niveaux du département (les seuls autorisés)
        filieres_dept = list(Filiere.objects.filter(departement=departement))
        if not filieres_dept:
            return Response({
                'error': f'Le département {departement.code} ne contient aucune filière. Créez d\'abord des filières.'
            }, status=400)

        filieres_map = {f.code.upper(): f for f in filieres_dept}
        niveaux_par_filiere = {
            f.id: {n.nom: n for n in Niveau.objects.filter(filiere=f)}
            for f in filieres_dept
        }

        created      = 0
        updated      = 0
        ens_crees    = 0
        errors       = []
        preview      = []
        filieres_vues = set()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            filiere_code = str(row[0]).strip().upper() if row[0] is not None else ''
            niveau_nom   = str(row[1]).strip().upper() if row[1] is not None else ''
            code_ue      = str(row[2]).strip()         if row[2] is not None else ''
            intitule     = str(row[3]).strip()         if row[3] is not None else ''
            enseignant_n = str(row[4]).strip()         if row[4] is not None else ''
            type_seance  = str(row[5]).strip().upper() if len(row) > 5 and row[5] is not None else 'CM'
            volume       = row[6] if len(row) > 6 else None

            # ── Validation bloquante ──────────────────────
            if not filiere_code:
                errors.append({'row': row_idx, 'code': '?', 'message': 'Code filière manquant.', 'blocking': True})
                continue
            if filiere_code not in filieres_map:
                errors.append({'row': row_idx, 'code': code_ue or '?',
                    'message': f'Filière "{filiere_code}" inconnue ou hors du département {departement.code}.',
                    'blocking': True})
                continue
            if not code_ue:
                errors.append({'row': row_idx, 'code': '?', 'message': 'Code UE manquant.', 'blocking': True})
                continue
            if not intitule:
                errors.append({'row': row_idx, 'code': code_ue, 'message': 'Intitulé manquant.', 'blocking': True})
                continue
            if type_seance not in TYPES_VALIDES:
                type_seance = 'CM'  # défaut tolérant
            if niveau_nom not in NIVEAUX_VALIDES:
                errors.append({'row': row_idx, 'code': code_ue, 'message': f'Niveau "{niveau_nom}" invalide (L1, L2, L3, M1 ou M2).', 'blocking': True})
                continue

            filiere = filieres_map[filiere_code]
            niveaux_map = niveaux_par_filiere[filiere.id]
            if niveau_nom not in niveaux_map:
                errors.append({'row': row_idx, 'code': code_ue,
                    'message': f'Le niveau {niveau_nom} n\'existe pas pour la filière {filiere.code}. Créez-le d\'abord.',
                    'blocking': True})
                continue

            niveau = niveaux_map[niveau_nom]
            filieres_vues.add(filiere_code)

            existante = Matiere.objects.filter(code=code_ue, niveau=niveau, session=session).first()

            # Enseignant : matcher par nom, créer si manquant
            enseignant = None
            if enseignant_n:
                if dry_run:
                    enseignant = Enseignant.objects.filter(nom__iexact=enseignant_n.split()[-1]).first()
                    if not enseignant:
                        errors.append({'row': row_idx, 'code': code_ue,
                            'message': f'Enseignant "{enseignant_n}" introuvable → sera créé automatiquement à l\'import.',
                            'blocking': False})
                else:
                    enseignant, was_new = _matcher_enseignant(enseignant_n, departement)
                    if was_new:
                        ens_crees += 1

            try:
                vol = float(volume) if volume else 2.5
            except (ValueError, TypeError):
                vol = 2.5

            action = 'update' if existante else 'create'
            preview.append({
                'row': row_idx, 'code': code_ue, 'intitule': intitule,
                'type_seance': type_seance, 'niveau': niveau_nom,
                'filiere': filiere_code,
                'enseignant': enseignant_n or '—',
                'volume_horaire': vol, 'action': action,
            })

            if not dry_run:
                _, created_now = Matiere.objects.update_or_create(
                    code=code_ue,
                    niveau=niveau,
                    session=session,
                    defaults={
                        'intitule':       intitule,
                        'type_seance':    type_seance,
                        'enseignant':     enseignant,
                        'volume_horaire': vol,
                    }
                )
                if created_now:
                    created += 1
                else:
                    updated += 1

        blocking_errors = [e for e in errors if e.get('blocking')]

        # Mise à jour du suivi : seulement si import réel (pas dry_run)
        # et si au moins une matière a été acceptée (pas que des erreurs).
        if not dry_run and (created + updated) > 0:
            suivi.date_import = timezone.now()
            suivi.fichier_nom = fichier.name
            suivi.nb_matieres = created + updated
            suivi.nb_erreurs  = len(blocking_errors)
            suivi.save()

        return Response({
            'dry_run':           dry_run,
            'session':           session.libelle,
            'session_id':        session.id,
            'departement':       departement.code,
            'departement_nom':   departement.nom,
            'filieres_touchees': sorted(filieres_vues),
            'created':           created,
            'updated':           updated,
            'enseignants_crees': ens_crees,
            'errors':            errors,
            'blocking_count':    len(blocking_errors),
            'preview':           preview if dry_run else [],
            'total_lignes':      len(preview) + len(blocking_errors),
        })
