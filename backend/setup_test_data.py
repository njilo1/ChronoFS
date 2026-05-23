"""
Script de création des données de test pour ChronoFS FS-UEB.
Crée les entités de base via l'API REST + génère les fichiers Excel d'import.

Usage (depuis backend/, venv activé) :
    python setup_test_data.py
"""
import os, json, requests, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL = "http://localhost:8000/api"

def post(endpoint, data):
    r = requests.post(f"{BASE_URL}/{endpoint}/", json=data)
    if r.status_code not in (200, 201):
        print(f"  ERREUR {endpoint}: {r.text[:120]}")
        return None
    return r.json()

def get_or_create(endpoint, data, lookup_field="code"):
    r = requests.get(f"{BASE_URL}/{endpoint}/?{lookup_field}={data[lookup_field]}&page_size=1")
    if r.ok:
        results = r.json().get("results", r.json() if isinstance(r.json(), list) else [])
        if results:
            print(f"  ↺ existe déjà : {data[lookup_field]}")
            return results[0]
    obj = post(endpoint, data)
    if obj:
        print(f"  ✓ créé : {data.get(lookup_field, data)}")
    return obj

# ── 1. Départements ───────────────────────────────────────────────────────────
print("\n=== Départements ===")
DEPTS = {}
for d in [
    {"nom": "Informatique et Technologies", "code": "INFO"},
    {"nom": "Sciences de la Vie",           "code": "BIOV"},
    {"nom": "Chimie",                       "code": "CHI"},
    {"nom": "Mathématiques-Physique",       "code": "MAPH"},
]:
    obj = get_or_create("departements", d)
    if obj: DEPTS[d["code"]] = obj["id"]

# ── 2. Campus (déjà créés par migration, on récupère leurs IDs) ───────────────
print("\n=== Campus ===")
CAMPUS = {}
r = requests.get(f"{BASE_URL}/campus/?page_size=20")
for c in r.json().get("results", []):
    CAMPUS[c["code"]] = c["id"]
    print(f"  ↺ {c['code']} — {c['nom']} ({c['ville']})")

# ── 3. Enseignants ────────────────────────────────────────────────────────────
print("\n=== Enseignants ===")
ENSEIGNANTS_DATA = [
    {"nom": "MBIDA",     "prenom": "Jean-Pierre", "grade": "Dr",  "email": "jmbida@ueb.cm",     "depts": ["INFO"]},
    {"nom": "NYABEYE",   "prenom": "Yannick",     "grade": "Dr",  "email": "ynyabeye@ueb.cm",   "depts": ["INFO"]},
    {"nom": "ESSOMBA",   "prenom": "Claude",      "grade": "Dr",  "email": "cessomba@ueb.cm",   "depts": ["INFO", "MAPH"]},
    {"nom": "ATANGANA",  "prenom": "Martin",      "grade": "Pr",  "email": "matangana@ueb.cm",  "depts": ["BIOV"]},
    {"nom": "FOUDA",     "prenom": "Marie-Claire","grade": "Dr",  "email": "mfouda@ueb.cm",     "depts": ["BIOV"]},
    {"nom": "NGONO",     "prenom": "Paul",        "grade": "Dr",  "email": "pngono@ueb.cm",     "depts": ["CHI"]},
    {"nom": "OWONA",     "prenom": "Robert",      "grade": "M",   "email": "rowona@ueb.cm",     "depts": ["INFO"]},
    {"nom": "MENYE",     "prenom": "Sylvie",      "grade": "Dr",  "email": "smenye@ueb.cm",     "depts": ["MAPH"]},
    {"nom": "BIKOK",     "prenom": "Honoré",      "grade": "Pr",  "email": "hbikok@ueb.cm",     "depts": ["CHI", "BIOV"]},
    {"nom": "NDOUMBE",   "prenom": "Alice",       "grade": "Dr",  "email": "andoumbe@ueb.cm",   "depts": ["INFO"]},
]
ENS = {}  # email → id
for e in ENSEIGNANTS_DATA:
    dept_ids = [DEPTS[k] for k in e["depts"] if k in DEPTS]
    obj = get_or_create("enseignants", {
        "nom": e["nom"], "prenom": e["prenom"], "grade": e["grade"],
        "email": e["email"], "departements": dept_ids, "est_actif": True,
    }, lookup_field="email")
    if obj: ENS[e["email"]] = obj["id"]

# ── 4. Filières ───────────────────────────────────────────────────────────────
print("\n=== Filières ===")
FILIERES_DATA = [
    # Ebolowa (CPF)
    {"nom": "Licence en Technologies de l'Information et de la Communication", "code": "TIC",  "effectif": 180, "dept": "INFO", "campus": "CPF"},
    {"nom": "Licence en Sciences de la Vie",                                   "code": "BIOV", "effectif": 150, "dept": "BIOV", "campus": "CPF"},
    {"nom": "Licence en Chimie",                                               "code": "CHI",  "effectif": 90,  "dept": "CHI",  "campus": "LYC"},
    {"nom": "Licence en Mathématiques-Physique",                               "code": "MAPH", "effectif": 110, "dept": "MAPH", "campus": "CPF"},
    # Monatélé (MON)
    {"nom": "Licence en Technologies de l'Information et de la Communication", "code": "TIC-MON", "effectif": 120, "dept": "INFO", "campus": "MON"},
    {"nom": "Licence en Sciences de la Vie",                                   "code": "BIO-MON", "effectif": 100, "dept": "BIOV", "campus": "MON"},
]
FILIERES = {}  # code → {id, niveaux: {nom→id}}
for f in FILIERES_DATA:
    campus_id = CAMPUS.get(f["campus"])
    obj = get_or_create("filieres", {
        "nom": f["nom"], "code": f["code"], "effectif": f["effectif"],
        "departement": DEPTS.get(f["dept"]), "campus": campus_id,
    })
    if obj:
        FILIERES[f["code"]] = {"id": obj["id"], "niveaux": {}}

# ── 5. Niveaux ────────────────────────────────────────────────────────────────
print("\n=== Niveaux ===")
NIVEAUX_PAR_FILIERE = {
    "TIC":     [("L1", 60), ("L2", 55), ("L3", 50)],
    "BIOV":    [("L1", 55), ("L2", 50), ("L3", 40)],
    "CHI":     [("L1", 35), ("L2", 30), ("L3", 25)],
    "MAPH":    [("L1", 45), ("L2", 40), ("L3", 35)],
    "TIC-MON": [("L1", 45), ("L2", 40), ("L3", 35)],
    "BIO-MON": [("L1", 40), ("L2", 35), ("L3", 30)],
}
for fil_code, niveaux in NIVEAUX_PAR_FILIERE.items():
    if fil_code not in FILIERES: continue
    fil_id = FILIERES[fil_code]["id"]
    for nom_niv, eff in niveaux:
        r = requests.get(f"{BASE_URL}/niveaux/?filiere={fil_id}&nom={nom_niv}")
        existing = r.json().get("results", []) if r.ok else []
        if existing:
            niv_id = existing[0]["id"]
            print(f"  ↺ {fil_code} {nom_niv}")
        else:
            obj = post("niveaux", {"nom": nom_niv, "effectif": eff, "filiere": fil_id})
            niv_id = obj["id"] if obj else None
            if niv_id: print(f"  ✓ {fil_code} {nom_niv} ({eff} ét.)")
        if niv_id:
            FILIERES[fil_code]["niveaux"][nom_niv] = niv_id

# ── 6. Salles supplémentaires ─────────────────────────────────────────────────
print("\n=== Salles ===")
SALLES_SUP = [
    {"nom": "Salle E", "capacite": 220, "campus": "CPF"},
    {"nom": "Salle A", "capacite": 100, "campus": "CPF"},
    {"nom": "Salle B", "capacite": 100, "campus": "CPF"},
    {"nom": "Salle M", "capacite": 40,  "campus": "CPF"},
    {"nom": "Salle N", "capacite": 40,  "campus": "CPF"},
    {"nom": "Salle D", "capacite": 40,  "campus": "LYC"},
    {"nom": "Salle F", "capacite": 20,  "campus": "CPF"},
    {"nom": "Salle G", "capacite": 20,  "campus": "CPF"},
    {"nom": "Amphi MON", "capacite": 150, "campus": "MON"},
    {"nom": "Salle M1",  "capacite": 50,  "campus": "MON"},
    {"nom": "Salle M2",  "capacite": 50,  "campus": "MON"},
]
for s in SALLES_SUP:
    get_or_create("salles", {"nom": s["nom"], "capacite": s["capacite"], "campus": CAMPUS.get(s["campus"]), "est_disponible": True}, lookup_field="nom")

# ── 7. Générer les fichiers Excel d'import ───────────────────────────────────
print("\n=== Génération des fichiers Excel ===")
os.makedirs("test_data", exist_ok=True)

# Définition des matières par filière et niveau
MATIERES = {
    "TIC": {
        "L1": [
            ("TIC111", "Algorithmique et Structures de Données", "CM", "jmbida@ueb.cm", 45),
            ("TIC112", "Algorithmique et Structures de Données", "TD", "jmbida@ueb.cm", 30),
            ("TIC113", "Introduction à la Programmation (Python)", "CM", "ynyabeye@ueb.cm", 30),
            ("TIC114", "Architecture des Ordinateurs",            "CM", "cessomba@ueb.cm", 30),
            ("TIC115", "Mathématiques Discrètes",                 "CM", "smenye@ueb.cm",   30),
            ("TIC116", "Anglais Technique",                       "CM", "rowona@ueb.cm",   20),
            ("TIC117", "Systèmes d'Exploitation (bases)",        "CM", "ynyabeye@ueb.cm", 25),
        ],
        "L2": [
            ("TIC221", "Programmation Orientée Objet (Java)",    "CM", "ynyabeye@ueb.cm", 30),
            ("TIC222", "Programmation Orientée Objet (Java)",    "TD", "ynyabeye@ueb.cm", 20),
            ("TIC223", "Bases de Données Relationnelles",         "CM", "jmbida@ueb.cm",   30),
            ("TIC224", "Bases de Données Relationnelles",         "TD", "jmbida@ueb.cm",   20),
            ("TIC225", "Réseaux Informatiques",                   "CM", "cessomba@ueb.cm", 30),
            ("TIC226", "Analyse et Conception (UML)",             "CM", "andoumbe@ueb.cm", 25),
            ("TIC227", "Probabilités et Statistiques",            "CM", "smenye@ueb.cm",   25),
        ],
        "L3": [
            ("TIC331", "Développement Web Full Stack",            "CM", "jmbida@ueb.cm",   30),
            ("TIC332", "Développement Web Full Stack",            "TP", "jmbida@ueb.cm",   20),
            ("TIC333", "Intelligence Artificielle",               "CM", "ynyabeye@ueb.cm", 30),
            ("TIC334", "Sécurité Informatique",                  "CM", "cessomba@ueb.cm", 25),
            ("TIC335", "Administration des Systèmes et Réseaux", "CM", "cessomba@ueb.cm", 25),
            ("TIC336", "Génie Logiciel",                         "CM", "andoumbe@ueb.cm", 25),
            ("TIC337", "Stage et Projet de Fin de Cycle",        "TP", "jmbida@ueb.cm",   60),
        ],
    },
    "BIOV": {
        "L1": [
            ("BIO111", "Biologie Cellulaire",                    "CM", "matangana@ueb.cm", 45),
            ("BIO112", "Biologie Cellulaire",                    "TD", "mfouda@ueb.cm",    20),
            ("BIO113", "Biochimie Générale",                     "CM", "hbikok@ueb.cm",    40),
            ("BIO114", "Zoologie des Invertébrés",               "CM", "matangana@ueb.cm", 30),
            ("BIO115", "Botanique Générale",                     "CM", "mfouda@ueb.cm",    30),
            ("BIO116", "Mathématiques et Statistiques Bio",      "CM", "smenye@ueb.cm",    25),
        ],
        "L2": [
            ("BIO221", "Génétique Générale",                     "CM", "matangana@ueb.cm", 40),
            ("BIO222", "Microbiologie",                          "CM", "mfouda@ueb.cm",    35),
            ("BIO223", "Physiologie Végétale",                   "CM", "mfouda@ueb.cm",    30),
            ("BIO224", "Immunologie",                            "CM", "matangana@ueb.cm", 25),
            ("BIO225", "Biochimie Métabolique",                  "CM", "hbikok@ueb.cm",    35),
            ("BIO226", "Écologie Générale",                      "CM", "matangana@ueb.cm", 25),
        ],
        "L3": [
            ("BIO331", "Biotechnologies",                        "CM", "mfouda@ueb.cm",    35),
            ("BIO332", "Parasitologie-Mycologie",                "CM", "matangana@ueb.cm", 30),
            ("BIO333", "Génétique Moléculaire",                  "CM", "matangana@ueb.cm", 35),
            ("BIO334", "Physiologie Animale",                    "CM", "hbikok@ueb.cm",    30),
            ("BIO335", "Projet de Recherche",                    "TP", "mfouda@ueb.cm",    50),
        ],
    },
    "TIC-MON": {
        "L1": [
            ("TIC111", "Algorithmique et Structures de Données", "CM", "andoumbe@ueb.cm", 45),
            ("TIC112", "Algorithmique et Structures de Données", "TD", "andoumbe@ueb.cm", 30),
            ("TIC113", "Introduction à la Programmation (Python)", "CM", "rowona@ueb.cm",  30),
            ("TIC115", "Mathématiques Discrètes",                 "CM", "smenye@ueb.cm",  30),
            ("TIC117", "Systèmes d'Exploitation (bases)",         "CM", "rowona@ueb.cm",  25),
        ],
        "L2": [
            ("TIC221", "Programmation Orientée Objet (Java)",    "CM", "andoumbe@ueb.cm", 30),
            ("TIC223", "Bases de Données Relationnelles",        "CM", "andoumbe@ueb.cm", 30),
            ("TIC225", "Réseaux Informatiques",                  "CM", "rowona@ueb.cm",   30),
            ("TIC226", "Analyse et Conception (UML)",            "CM", "andoumbe@ueb.cm", 25),
        ],
        "L3": [
            ("TIC331", "Développement Web Full Stack",           "CM", "andoumbe@ueb.cm", 30),
            ("TIC334", "Sécurité Informatique",                 "CM", "rowona@ueb.cm",   25),
            ("TIC336", "Génie Logiciel",                        "CM", "andoumbe@ueb.cm", 25),
        ],
    },
}

C_HEADER  = "0D1520"
C_GOLD    = "C9A450"
C_EXEMPLE = "1A2A3A"
C_ODD     = "0A0F16"
C_EVEN    = "080D14"

def make_excel(fil_code, matieres_par_niveau):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matières"

    # En-tête
    headers = ["code", "intitule", "type_seance (CM/TD/TP)", "niveau (L1-M2)", "enseignant_email", "volume_horaire (h)"]
    widths  = [14, 48, 22, 18, 30, 20]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color=C_GOLD, size=11)
        c.fill      = PatternFill(start_color=C_HEADER, end_color=C_HEADER, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    row = 2
    for niveau, matieres in matieres_par_niveau.items():
        for i, (code, intitule, type_s, email, vol) in enumerate(matieres):
            bg = C_ODD if (row % 2 == 0) else C_EVEN
            for col, val in enumerate([code, intitule, type_s, niveau, email, vol], 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font      = Font(color="D0C8B8", size=10)
                c.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                c.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 17
            row += 1

    path = f"test_data/import_{fil_code.lower()}.xlsx"
    wb.save(path)
    print(f"  ✓ {path}  ({row - 2} lignes)")
    return path

for fil_code, niveaux_data in MATIERES.items():
    make_excel(fil_code, niveaux_data)

# ── Résumé ────────────────────────────────────────────────────────────────────
print("\n" + "="*60)
print("DONNÉES DE TEST CRÉÉES")
print("="*60)
print(f"  Départements  : {len(DEPTS)}")
print(f"  Campus        : {len(CAMPUS)}")
print(f"  Enseignants   : {len(ENS)}")
print(f"  Filières      : {len(FILIERES)}")
print(f"  Fichiers Excel: test_data/  ({len(MATIERES)} fichiers)")
print()
print("COMPTES ENSEIGNANTS (à utiliser dans l'import Excel) :")
for e in ENSEIGNANTS_DATA:
    print(f"  {e['email']:<30} {e['grade']}. {e['nom']} {e['prenom']}")
print()
print("PROCHAINES ÉTAPES :")
print("  1. Aller dans /import → sélectionner une filière → importer le fichier Excel correspondant")
print("  2. Aller dans /plannings → créer une semaine → générer l'EDT")
print("  3. Cliquer sur la grille ou exporter en PDF")
